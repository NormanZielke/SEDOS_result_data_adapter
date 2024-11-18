
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from functions import name_function, var_name_function

# Vorbereitung Output

# create dataframe for results
columns = ["scenario", "parameter", "process", "sector", "category",
               "specification", "groups", "new", "input_groups",
               "output_groups", "unit", "value"]

output = pd.DataFrame(columns= columns)

# Vorbereitung Input

data = pd.read_csv("results.csv",
                        sep=",")

# Wendet eine String-Ersetzung auf jeden Wert in der Spalte name an.
data['name'] = data['name'].str.replace(r'--\d+$', '', regex=True)

# Entfernen aller Zeilen, bei denen die Eintr채ge in der Spalte "name" mit "helper" beginnen
data = data[~data['name'].str.startswith('helper')]
data.reset_index(drop=True)
# --------------------------------------------------------------------------------------------------------------------->

# Erzeugen von zwei test Zeilen in data f체r var_name == invest
test_data = {"name": ["ind_steel_casting_1","ind_steel_casting_2"],
        "var_name": ["invest_out_exo_steel","invest_out_exo_steel"],
        "var_value": [999,888],
        "region": ["DE","DE"],
        "type":[{},{}],
        "carrier": [{},{}],
        "tech":["tech","tech"],
        "unit":["MWh","kg"]
}
df_test = pd.DataFrame(test_data)

data['unit'] = "MWh"

data2 = pd.concat([data, df_test], ignore_index=True)


# --------------------------------------------------------------------------------------------------------------------->
'''
# fill the results dataframe with data

name_function(data,output,columns) # evtl. columns als argument entfernen

var_name_function(data,output)

output.value = data.var_value

output.scenario = "o_steel_tokio"
'''
# --------------------------------------------------------------------------------------------------------------------->

# fill the results dataframe with data incl. testdata

name_function(data2,output,columns) # evtl. columns als argument entfernen

var_name_function(data2,output)

output.value = data2.var_value

output.scenario = "o_steel_tokio"

# --------------------------------------------------------------------------------------------------------------------->

# save data as excel sheet

with pd.ExcelWriter("SEDOS_output.xlsx") as writer:
    output.to_excel(writer, sheet_name="SEDOS_output", index=False)

# Lade die erstellte Excel-Datei mit openpyxl
wb = load_workbook("SEDOS_output.xlsx")

# Funktion zum Anpassen der Spaltenbreite
def adjust_column_width(sheet):
    for col_index, column_cells in enumerate(sheet.columns, start=1):
        # Berechne die maximale L채nge der Werte in der Spalte
        max_length = max(len(str(cell.value) or "") for cell in column_cells)  # Handle None values
        column_letter = get_column_letter(col_index)
        # Setze die Spaltenbreite basierend auf der maximalen L채nge der Zelleninhalte
        sheet.column_dimensions[column_letter].width = max_length + 5

    # Passe die Spaltenbreite an
sheet = wb["SEDOS_output"]
adjust_column_width(sheet)

